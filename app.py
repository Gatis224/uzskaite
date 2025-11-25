# app.py
import re
import calendar
import io
from datetime import date
from flask import Flask, request, send_file, render_template_string
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import holidays
from openpyxl.utils import get_column_letter

app = Flask(__name__)

LATV_MONTHS = {
    1: "JANVARIS", 2: "FEBRUARIS", 3: "MARTS", 4: "APRILIS",
    5: "MAIJS", 6: "JUNIJS", 7: "JULIJS", 8: "AUGUSTS",
    9: "SEPTEMBRIS", 10: "OKTOBRIS", 11: "NOVEMBRIS", 12: "DECEMBRIS"
}
LATV_MONTHS_LC = {
    1: "janvāris", 2: "februāris", 3: "marts", 4: "aprīlis",
    5: "maijs", 6: "jūnijs", 7: "jūlijs", 8: "augusts",
    9: "septembris", 10: "oktobris", 11: "novembris", 12: "decembris"
}

GRAY_FILL = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)

def next_month(year, month):
    return (year + 1, 1) if month == 12 else (year, month + 1)

def find_header_cell(ws):
    pattern = re.compile(r"(\d{4})\.(\w+)", re.IGNORECASE)
    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                m = pattern.search(v.strip())
                if m:
                    return r, c, v.strip()
    raise ValueError("Nevar atrast virsrakstu, kur ir 'YYYY.month'")

def find_day_row(ws):
    for r in range(1, 40):
        vals = [str(ws.cell(row=r, column=c).value or "").strip()
                for c in range(1, ws.max_column + 1)]
        if "1" in vals and "2" in vals and "3" in vals:
            return r
    raise ValueError("Nevar atrast dienu rindu.")

def find_workers(ws, start_row):
    workers = []
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            break
        # Regex tagad pieņem gan "5." gan "5"
        if re.match(r"^\d+\.?$", str(v).strip()):
            workers.append(r)
        else:
            break
    return workers

def process_workbook(stream):
    wb = load_workbook(stream)
    ws = wb.active

    # =======================
    # 1) Atrodam mēnesi un gadu
    # =======================
    h_row, h_col, h_text = find_header_cell(ws)
    m = re.search(r"(\d{4})\.(\w+)", h_text)
    year = int(m.group(1))
    mstr = m.group(2).lower()

    month_map = {
        'janvaris':1,'janvāris':1,'februaris':2, 'februāris':2,'marts':3,'aprilis':4, 'aprīlis':4,
        'maijs':5,'jūnijs':6,'junijs':6,'jūlijs':7,'julijs':7,
        'augusts':8,'septembris':9,'oktobris':10,'novembris':11,'decembris':12
    }
    if mstr not in month_map:
        raise ValueError("Neatpazīts mēnesis: " + mstr)

    month = month_map[mstr]
    ny, nm = next_month(year, month)

    # Jaunais virsraksts
    ws.cell(row=h_row, column=h_col).value = f"{ny}.{LATV_MONTHS_LC[nm]}"
    ws.cell(row=h_row, column=h_col).font = Font(size=14, bold=False)

    # =======================
    # 2) Atrodam dienu rindu un dienu sākuma kolonnu
    # =======================
    day_row = find_day_row(ws)

    start_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=day_row, column=c).value
        if v and str(v).strip() == "1":
            start_col = c
            break

    if start_col is None:
        raise ValueError("Nevar atrast kolonu ar '1' dienu.")

    # =======================
    # 3) Atrodam darbinieku rindas
    # =======================
    workers = find_workers(ws, day_row + 1)
    if not workers:
        raise ValueError("Nevar atrast nevienu darbinieka rindu.")

    first_worker = workers[0]
    last_worker = workers[-1]

    # =======================
    # 4) NORMALIZĒ VISAS 31 DIENAS
    # =======================
    # Vispirms notīra VISUS esošos datumus, lai neatstātu nejaušus formatējumus


    for d in range(1, 32):
        col = start_col + (d - 1)
        # Notīrām dienu galveni
        ws.cell(row=day_row, column=col).value = None
        ws.cell(row=day_row, column=col).fill = CLEAR_FILL
        ws.cell(row=day_row, column=col).font = Font(size=8)

        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].hidden = False
        ws.column_dimensions[col_letter].width = 3.0

        # Notīrām katru darbinieka šūnu
        for r in workers:
            cell = ws.cell(row=r, column=col)
            cell.value = None
            cell.fill = CLEAR_FILL
            cell.font = Font(size=8)

    # =======================
    # 5) Ievieto jauno mēnesi (1–31)
    # =======================
    month_len = calendar.monthrange(ny, nm)[1]
    lv_holidays = holidays.CountryHoliday("LV", years=[ny])

    for d in range(1, 32):
        col = start_col + (d - 1)
        hdr = ws.cell(row=day_row, column=col)

        if d <= month_len:
            # Dienas numurs
            hdr.value = str(d)
            hdr.font = Font(size=8)
            hdr.fill = CLEAR_FILL

            # Weekends & svētki
            dt = date(ny, nm, d)
            if dt.weekday() >= 5 or dt in lv_holidays:
                hdr.fill = GRAY_FILL
        else:
            # Neeksistējošu dienu galvene
            hdr.value = None
            hdr.fill = CLEAR_FILL

    # =======================
    # 6) Aizpilda D/E darbiniekiem (1–31)
    # =======================
    for d in range(1, 32):
        col = start_col + (d - 1)

        if d <= month_len:
            dt = date(ny, nm, d)
            is_weekend = dt.weekday() >= 5 or dt in lv_holidays
            for r in workers:
                cell = ws.cell(row=r, column=col)
                if is_weekend:
                    cell.value = None
                    cell.fill = GRAY_FILL
                else:
                    # Pirmdiena–Ceturtdiena = D, Piektdiena = E
                    # Pirmssvētku diena = "F"
                    next_day = dt.toordinal() + 1  # nākamās dienas numurs
                    next_dt = date.fromordinal(next_day)

                    if next_dt in lv_holidays:
                        cell.value = "F"
                    else:
                        # Normāla darba diena
                        cell.value = "D" if dt.weekday() <= 3 else "E"
                    cell.fill = CLEAR_FILL
                cell.font = Font(size=8)
        else:
            # Neeksistējošās dienas
            for r in workers:
                ws.cell(row=r, column=col).value = None
                ws.cell(row=r, column=col).fill = CLEAR_FILL
                ws.cell(row=r, column=col).font = Font(size=8)

    # =======================
    # 7) Izdzēš liekās kolonnas (ja < 31)
    # =======================
    if month_len < 31:
        first_extra_col = start_col + month_len
        last_extra_col = start_col + 30

        for col in range(first_extra_col, last_extra_col + 1):
            col_letter = get_column_letter(col)

            # Slēpjam kolonnu
            ws.column_dimensions[col_letter].hidden = True

            # Attīram galveni
            cell = ws.cell(row=day_row, column=col)
            cell.value = None
            cell.fill = CLEAR_FILL
            cell.font = Font(size=8)

            # Attīram darbinieku šūnas
            for r in workers:
                c = ws.cell(row=r, column=col)
                c.value = None
                c.fill = CLEAR_FILL
                c.font = Font(size=8)
    # =======================
    # 8) Saglabā
    # =======================
    outname = f"{LATV_MONTHS[nm]}_Kanceleja_{ny}.xlsx"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, outname


# =========================
# Flask UI
# =========================

HTML = """
<!doctype html>
<html lang="lv">
<head>
<meta charset="UTF-8">
<title>Excel ģenerators – nākamais mēnesis</title>
<style>
    body {
        font-family: Arial, sans-serif;
        background: #f2f4f7;
        padding: 40px;
        display: flex;
        justify-content: center;
    }
    .container {
        max-width: 480px;
        width: 100%;
        background: white;
        padding: 30px;
        border-radius: 12px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        text-align: center;
    }
    h2 {
        font-weight: 600;
        margin-bottom: 20px;
    }
    .upload-box {
        border: 2px dashed #6c8df5;
        padding: 20px;
        border-radius: 10px;
        background: #f8faff;
        cursor: pointer;
        color: #6c7a89;
        transition: 0.2s;
    }
    .upload-box:hover {
        border-color: #456df0;
        background: #f1f5ff;
    }
    #fileInput {
        display: none;
    }
    .file-name {
        margin-top: 10px;
        font-style: italic;
        color: #333;
    }
    button {
        margin-top: 20px;
        padding: 12px 30px;
        background: #456df0;
        border: none;
        border-radius: 8px;
        font-size: 16px;
        color: white;
        cursor: pointer;
        transition: 0.2s;
    }
    button:hover {
        background: #2f4ec7;
    }
    p.note {
        margin-top: 15px;
        color: #666;
        font-size: 13px;
    }
    form {
        height: 200px;
        display: flex;
        flex-direction: column;
        gap: 8px; /* attālums starp elementiem */
    }

        input[type="file"] {
        max-width: 300px;
    }
    .note {
        font-size: 0.9em;
        color: #666;
    }

</style>
</head>
<body>

<div class="container">
    <h2>Ģenerēt nākamo mēnesi</h2>

    <form action="/" method="post" enctype="multipart/form-data">

        <label for="fileInput" class="upload-box" id="dropArea">
            Klikšķini vai ievelc Excel failu (.xlsx)
        </label>

        <input type="file" id="fileInput" name="file" accept=".xlsx">

        <div class="file-name" id="fileName"></div>

        <button type="submit">Ģenerēt</button>
    </form>
</div>

<script>
// Faila nosaukuma parādīšana
const fileInput = document.getElementById("fileInput");
const fileName = document.getElementById("fileName");

fileInput.addEventListener("change", () => {
    if (fileInput.files.length > 0) {
        fileName.textContent = "Izvēlēts fails: " + fileInput.files[0].name;
    }
});

// Drag & Drop
const dropArea = document.getElementById("dropArea");

dropArea.addEventListener("dragover", (e) => {
    e.preventDefault();
    dropArea.style.background = "#e8edff";
});

dropArea.addEventListener("dragleave", () => {
    dropArea.style.background = "#f8faff";
});

dropArea.addEventListener("drop", (e) => {
    e.preventDefault();
    dropArea.style.background = "#f8faff";

    const file = e.dataTransfer.files[0];

    // Atļaujam tikai .xlsx
    if (file && file.name.toLowerCase().endsWith(".xlsx")) {
        fileInput.files = e.dataTransfer.files;
        fileName.textContent = "Izvēlēts fails: " + file.name;
    } else {
        fileName.textContent = "❌ Atļauti tikai .xlsx faili";
    }
});
</script>

</body>
</html>
"""

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        f = request.files.get("file")
        if not f:
            return "Nav augšupielādēta faila", 400
        try:
            data, filename = process_workbook(f.stream)
        except Exception as e:
            return f"Kļūda apstrādājot failu: {e}", 500

        return send_file(
            data,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return render_template_string(HTML)

if __name__ == "__main__":
    app.run()
